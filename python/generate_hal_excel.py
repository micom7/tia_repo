#!/usr/bin/env python3
"""
Генератор HAL коду для TIA Portal з Excel конфігурації
Версія: 2.0 (Excel підтримка)
Автор: Elevator Automation Team

Підтримує:
- .xlsx (Office 2007+) - через openpyxl
- .xls (Office 97-2003) - через xlrd
- .csv - через csv module

Залежності:
    pip install openpyxl xlrd pandas

Використання:
    python generate_hal_excel.py IO_Config.xlsx
    python generate_hal_excel.py IO_Config.xls
    python generate_hal_excel.py IO_Config.csv
"""

import os
import sys
from datetime import datetime
from typing import List, Dict, Optional
import argparse


class ExcelReader:
    """Універсальний читач Excel/CSV файлів"""
    
    @staticmethod
    def read_excel_openpyxl(filepath: str) -> List[Dict]:
        """Читання .xlsx через openpyxl"""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("Встановіть openpyxl: pip install openpyxl")
        
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        # Читання заголовків
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Читання даних
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:  # пропустити порожні рядки
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                value = row[i] if i < len(row) else None
                # Конвертувати у string
                if value is None:
                    row_dict[header] = ''
                else:
                    row_dict[header] = str(value).strip()
            data.append(row_dict)
        
        return data
    
    @staticmethod
    def read_excel_xlrd(filepath: str) -> List[Dict]:
        """Читання .xls через xlrd"""
        try:
            import xlrd
        except ImportError:
            raise ImportError("Встановіть xlrd: pip install xlrd")
        
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        
        # Читання заголовків
        headers = ws.row_values(0)
        
        # Читання даних
        data = []
        for row_idx in range(1, ws.nrows):
            row_values = ws.row_values(row_idx)
            if not row_values[0]:  # пропустити порожні рядки
                continue
            row_dict = {}
            for i, header in enumerate(headers):
                value = row_values[i] if i < len(row_values) else ''
                # Конвертувати у string
                if isinstance(value, float):
                    # Якщо це ціле число, видалити .0
                    if value.is_integer():
                        row_dict[header] = str(int(value))
                    else:
                        row_dict[header] = str(value)
                else:
                    row_dict[header] = str(value).strip()
            data.append(row_dict)
        
        return data
    
    @staticmethod
    def read_csv(filepath: str) -> List[Dict]:
        """Читання CSV"""
        import csv
        
        data = []
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                # Конвертувати всі значення у string та прибрати пробіли
                row_dict = {k: str(v).strip() for k, v in row.items()}
                data.append(row_dict)
        
        return data
    
    @staticmethod
    def read(filepath: str) -> List[Dict]:
        """Автоматичне визначення формату та читання"""
        ext = os.path.splitext(filepath)[1].lower()
        
        if ext == '.xlsx':
            return ExcelReader.read_excel_openpyxl(filepath)
        elif ext == '.xls':
            return ExcelReader.read_excel_xlrd(filepath)
        elif ext == '.csv':
            return ExcelReader.read_csv(filepath)
        else:
            raise ValueError(f"Непідтримуваний формат: {ext}. Використовуйте .xlsx, .xls або .csv")


class HALGenerator:
    """Генератор SCL коду для HAL"""
    
    # Очікувані колонки
    REQUIRED_COLUMNS = [
        'Slot', 'DeviceType', 'TypedIndex', 'Name',
        'DI_Speed_Addr', 'DI_Breaker_Addr', 'DI_Overflow_Addr', 'DO_Run_Addr',
        'DI_Speed_Invert', 'DI_Breaker_Invert', 'DI_Overflow_Invert', 'DO_Run_Invert',
        'Enable_OK'
    ]
    
    OPTIONAL_COLUMNS = ['Comment', 'Location', 'Description']
    
    def __init__(self, filepath: str):
        self.filepath = filepath
        self.mechanisms = []
        self.load_config()
    
    def load_config(self):
        """Завантажити конфігурацію з файлу"""
        print(f"📂 Читання файлу: {self.filepath}")
        
        try:
            data = ExcelReader.read(self.filepath)
        except Exception as e:
            print(f"❌ Помилка читання файлу: {e}")
            sys.exit(1)
        
        # Перевірка обов'язкових колонок
        if not data:
            print("❌ Файл порожній")
            sys.exit(1)
        
        first_row = data[0]
        missing_cols = [col for col in self.REQUIRED_COLUMNS if col not in first_row]
        
        if missing_cols:
            print(f"❌ Відсутні обов'язкові колонки: {', '.join(missing_cols)}")
            print(f"   Наявні колонки: {', '.join(first_row.keys())}")
            sys.exit(1)
        
        # Фільтр: тільки редлери (DeviceType=2) і активні (Enable_OK=TRUE)
        for row in data:
            if row['DeviceType'] == '2' and row['Enable_OK'].upper() in ('TRUE', '1', 'YES'):
                self.mechanisms.append(row)
        
        print(f"✓ Завантажено {len(self.mechanisms)} активних механізмів")
        
        if not self.mechanisms:
            print("⚠️  Увага: не знайдено жодного активного редлера (DeviceType=2, Enable_OK=TRUE)")
    
    def validate_config(self) -> List[str]:
        """Валідація конфігурації"""
        errors = []
        
        # Перевірка дублікатів Slot
        slots = [m['Slot'] for m in self.mechanisms]
        duplicates = [s for s in slots if slots.count(s) > 1]
        if duplicates:
            errors.append(f"Дублікати Slot: {set(duplicates)}")
        
        # Перевірка дублікатів TypedIndex
        indices = [m['TypedIndex'] for m in self.mechanisms]
        duplicates = [i for i in indices if indices.count(i) > 1]
        if duplicates:
            errors.append(f"Дублікати TypedIndex: {set(duplicates)}")
        
        # Перевірка дублікатів адрес
        used_inputs = {}
        used_outputs = {}
        
        for mech in self.mechanisms:
            name = mech['Name']
            
            # Входи
            for addr_key in ['DI_Speed_Addr', 'DI_Breaker_Addr', 'DI_Overflow_Addr']:
                addr = mech[addr_key]
                if addr in used_inputs:
                    errors.append(f"Дублікат входу {addr}: {name} та {used_inputs[addr]}")
                used_inputs[addr] = name
            
            # Виходи
            addr = mech['DO_Run_Addr']
            if addr in used_outputs:
                errors.append(f"Дублікат виходу {addr}: {name} та {used_outputs[addr]}")
            used_outputs[addr] = name
            
            # Перевірка формату адреси
            for addr_key in ['DI_Speed_Addr', 'DI_Breaker_Addr', 'DI_Overflow_Addr']:
                addr = mech[addr_key]
                if not addr.startswith('%I'):
                    errors.append(f"Невалідна адреса входу {addr} у {name} (має починатися з %I)")
            
            addr = mech['DO_Run_Addr']
            if not addr.startswith('%Q'):
                errors.append(f"Невалідна адреса виходу {addr} у {name} (має починатися з %Q)")
        
        # Перевірка TypedIndex без пропусків
        indices_int = sorted([int(m['TypedIndex']) for m in self.mechanisms])
        for i, idx in enumerate(indices_int):
            if idx != i:
                errors.append(f"Пропуск у TypedIndex: очікується {i}, знайдено {idx}")
                break
        
        return errors
    
    def generate_header(self, function_name: str, description: str) -> List[str]:
        """Згенерувати заголовок функції"""
        lines = []
        lines.append(f"// {'='*78}")
        lines.append(f"// {function_name}")
        lines.append(f"// {description}")
        lines.append(f"//")
        lines.append(f"// Згенеровано автоматично: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"// Джерело: {os.path.basename(self.filepath)}")
        lines.append(f"// Механізмів: {len(self.mechanisms)}")
        lines.append(f"// {'='*78}")
        lines.append("")
        return lines
    
    def generate_hal_read(self, output_file: str):
        """Згенерувати FC_HAL_Read_Redler_Static"""
        lines = []
        
        # Header
        lines.extend(self.generate_header(
            "FC_HAL_Read_Redler_Static",
            "HAL Read Phase: Читання входів та нормалізація сигналів"
        ))
        
        # Function declaration
        lines.append('FUNCTION "FC_HAL_Read_Redler_Static" : VOID')
        lines.append("{ S7_Optimized_Access := 'TRUE' }")
        lines.append("")
        
        # VAR_IN_OUT
        lines.append("VAR_IN_OUT")
        lines.append("    Redler : ARRAY[*] OF \"UDT_Redler\";")
        lines.append("END_VAR")
        lines.append("")
        
        # VAR_TEMP - AT variables
        lines.append("VAR_TEMP")
        lines.append("    // Прямий доступ до входів через AT")
        lines.append("")
        
        for mech in self.mechanisms:
            idx = mech['TypedIndex']
            name = mech['Name']
            slot = mech['Slot']
            lines.append(f"    // [{idx}] {name} (Slot {slot})")
            lines.append(f"    R{idx}_Speed    AT {mech['DI_Speed_Addr']} : BOOL;")
            lines.append(f"    R{idx}_Breaker  AT {mech['DI_Breaker_Addr']} : BOOL;")
            lines.append(f"    R{idx}_Overflow AT {mech['DI_Overflow_Addr']} : BOOL;")
            lines.append("")
        
        lines.append("END_VAR")
        lines.append("")
        
        # BEGIN
        lines.append("BEGIN")
        lines.append("    // " + "="*74)
        lines.append("    // ЧИТАННЯ ТА НОРМАЛІЗАЦІЯ ВХОДІВ")
        lines.append("    // " + "="*74)
        lines.append("")
        
        for mech in self.mechanisms:
            idx = mech['TypedIndex']
            name = mech['Name']
            slot = mech['Slot']
            comment = mech.get('Comment', '')
            
            lines.append(f"    // " + "-"*74)
            lines.append(f"    // [{idx}] {name} (Slot {slot})")
            if comment:
                lines.append(f"    // {comment}")
            lines.append(f"    // " + "-"*74)
            
            # DI_Speed
            invert = mech['DI_Speed_Invert'].upper() in ('TRUE', '1', 'YES')
            if invert:
                lines.append(f"    Redler[{idx}].DI_Speed_OK := NOT R{idx}_Speed;")
            else:
                lines.append(f"    Redler[{idx}].DI_Speed_OK := R{idx}_Speed;")
            
            # DI_Breaker
            invert = mech['DI_Breaker_Invert'].upper() in ('TRUE', '1', 'YES')
            if invert:
                lines.append(f"    Redler[{idx}].DI_Breaker_OK := NOT R{idx}_Breaker;")
            else:
                lines.append(f"    Redler[{idx}].DI_Breaker_OK := R{idx}_Breaker;")
            
            # DI_Overflow
            invert = mech['DI_Overflow_Invert'].upper() in ('TRUE', '1', 'YES')
            if invert:
                lines.append(f"    Redler[{idx}].DI_Overflow_OK := NOT R{idx}_Overflow;")
            else:
                lines.append(f"    Redler[{idx}].DI_Overflow_OK := R{idx}_Overflow;")
            
            lines.append("")
        
        lines.append("END_FUNCTION")
        
        # Записати файл
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✓ Згенеровано: {output_file}")
        print(f"  - Функція: FC_HAL_Read_Redler_Static")
        print(f"  - Рядків коду: {len(lines)}")
        print(f"  - Входів: {len(self.mechanisms) * 3}")
    
    def generate_hal_write(self, output_file: str):
        """Згенерувати FC_HAL_Write_Redler_Static"""
        lines = []
        
        # Header
        lines.extend(self.generate_header(
            "FC_HAL_Write_Redler_Static",
            "HAL Write Phase: Запис виходів з SafeMode"
        ))
        
        # Function declaration
        lines.append('FUNCTION "FC_HAL_Write_Redler_Static" : VOID')
        lines.append("{ S7_Optimized_Access := 'TRUE' }")
        lines.append("")
        
        # VAR_INPUT
        lines.append("VAR_INPUT")
        lines.append("    SafeMode : BOOL := FALSE;  // TRUE = всі виходи OFF (аварійна зупинка)")
        lines.append("END_VAR")
        lines.append("")
        
        # VAR_IN_OUT
        lines.append("VAR_IN_OUT")
        lines.append("    Redler : ARRAY[*] OF \"UDT_Redler\";")
        lines.append("END_VAR")
        lines.append("")
        
        # VAR_TEMP
        lines.append("VAR_TEMP")
        lines.append("    // Прямий доступ до виходів через AT")
        lines.append("")
        
        for mech in self.mechanisms:
            idx = mech['TypedIndex']
            name = mech['Name']
            slot = mech['Slot']
            lines.append(f"    // [{idx}] {name} (Slot {slot})")
            lines.append(f"    R{idx}_Run AT {mech['DO_Run_Addr']} : BOOL;")
            lines.append(f"    R{idx}_Cmd : BOOL;  // проміжна команда")
            lines.append("")
        
        lines.append("END_VAR")
        lines.append("")
        
        # BEGIN
        lines.append("BEGIN")
        lines.append("    // " + "="*74)
        lines.append("    // ЗАПИС ВИХОДІВ З SAFEGUARD")
        lines.append("    // " + "="*74)
        lines.append("")
        
        for mech in self.mechanisms:
            idx = mech['TypedIndex']
            name = mech['Name']
            slot = mech['Slot']
            comment = mech.get('Comment', '')
            
            lines.append(f"    // " + "-"*74)
            lines.append(f"    // [{idx}] {name} (Slot {slot})")
            if comment:
                lines.append(f"    // {comment}")
            lines.append(f"    // " + "-"*74)
            
            # Отримати команду від механізму з урахуванням SafeMode
            lines.append(f"    R{idx}_Cmd := Redler[{idx}].DO_Run AND NOT SafeMode;")
            
            # Записати у вихід з урахуванням інверсії
            invert = mech['DO_Run_Invert'].upper() in ('TRUE', '1', 'YES')
            if invert:
                lines.append(f"    R{idx}_Run := NOT R{idx}_Cmd;")
            else:
                lines.append(f"    R{idx}_Run := R{idx}_Cmd;")
            
            lines.append("")
        
        lines.append("END_FUNCTION")
        
        # Записати файл
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✓ Згенеровано: {output_file}")
        print(f"  - Функція: FC_HAL_Write_Redler_Static")
        print(f"  - Рядків коду: {len(lines)}")
        print(f"  - Виходів: {len(self.mechanisms)}")
    
    def generate_db_io_config(self, output_file: str):
        """Згенерувати DB_IO_Config_Doc з даними"""
        lines = []
        
        # Header
        lines.extend(self.generate_header(
            "DB_IO_Config_Doc",
            "Документація конфігурації I/O мапінгу"
        ))
        
        lines.append('DATA_BLOCK "DB_IO_Config_Doc"')
        lines.append("{ S7_Optimized_Access := 'TRUE' }")
        lines.append("")
        lines.append("VAR")
        lines.append("    // Налаштування HAL")
        lines.append("    HAL_Enabled    : BOOL := TRUE;   // увімкнути HAL")
        lines.append("    HAL_SafeMode   : BOOL := FALSE;  // safe mode = всі виходи OFF")
        lines.append("    HAL_DiagMode   : BOOL := FALSE;  // діагностичний режим")
        lines.append("")
        lines.append("    // Статистика")
        lines.append("    HAL_ReadCycles  : UDINT;  // лічильник циклів READ")
        lines.append("    HAL_WriteCycles : UDINT;  // лічильник циклів WRITE")
        lines.append("    HAL_Errors      : UINT;   // лічильник помилок")
        lines.append("")
        lines.append("    // Конфігурація проекту")
        lines.append("    TotalMechanisms : UINT := " + str(len(self.mechanisms)) + ";")
        lines.append("    ConfigSource    : STRING[64] := '" + os.path.basename(self.filepath) + "';")
        lines.append("END_VAR")
        lines.append("")
        lines.append("BEGIN")
        lines.append("    // " + "="*74)
        lines.append("    // МАПІНГ МЕХАНІЗМІВ (для документації)")
        lines.append("    // " + "="*74)
        
        for mech in self.mechanisms:
            idx = mech['TypedIndex']
            slot = mech['Slot']
            name = mech['Name']
            comment = mech.get('Comment', '')
            
            lines.append(f"    //")
            lines.append(f"    // Slot {slot}: TypedIndex {idx}")
            lines.append(f"    // Name: {name}")
            if comment:
                lines.append(f"    // Comment: {comment}")
            lines.append(f"    // Inputs:  {mech['DI_Speed_Addr']}, {mech['DI_Breaker_Addr']}, {mech['DI_Overflow_Addr']}")
            lines.append(f"    // Output:  {mech['DO_Run_Addr']}")
            
            # Інверсії
            inv = []
            if mech['DI_Speed_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('Speed')
            if mech['DI_Breaker_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('Breaker')
            if mech['DI_Overflow_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('Overflow')
            if mech['DO_Run_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('Run')
            
            if inv:
                lines.append(f"    // Inverted: {', '.join(inv)}")
        
        lines.append("")
        lines.append("END_DATA_BLOCK")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✓ Згенеровано: {output_file}")
    
    def generate_documentation(self, output_file: str):
        """Згенерувати документацію I/O мапінгу у Markdown"""
        lines = []
        lines.append("# HAL I/O Configuration")
        lines.append("")
        lines.append(f"**Згенеровано:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"**Джерело:** {self.filepath}")
        lines.append(f"**Механізмів:** {len(self.mechanisms)}")
        lines.append("")
        lines.append("## Таблиця мапінгу")
        lines.append("")
        lines.append("| Slot | Type | Idx | Name | Speed | Breaker | Overflow | Run | Inv | Comment |")
        lines.append("|------|------|-----|------|-------|---------|----------|-----|-----|---------|")
        
        for mech in self.mechanisms:
            inv = []
            if mech['DI_Speed_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('S')
            if mech['DI_Breaker_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('B')
            if mech['DI_Overflow_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('O')
            if mech['DO_Run_Invert'].upper() in ('TRUE', '1', 'YES'):
                inv.append('R')
            
            inv_str = ','.join(inv) if inv else '-'
            
            lines.append(
                f"| {mech['Slot']} "
                f"| {mech['DeviceType']} "
                f"| {mech['TypedIndex']} "
                f"| {mech['Name']} "
                f"| {mech['DI_Speed_Addr']} "
                f"| {mech['DI_Breaker_Addr']} "
                f"| {mech['DI_Overflow_Addr']} "
                f"| {mech['DO_Run_Addr']} "
                f"| {inv_str} "
                f"| {mech.get('Comment', '')} |"
            )
        
        lines.append("")
        lines.append("**Легенда інверсії:** S=Speed, B=Breaker, O=Overflow, R=Run")
        lines.append("")
        lines.append("## Статистика")
        lines.append("")
        lines.append(f"- **Всього входів:** {len(self.mechanisms) * 3}")
        lines.append(f"- **Всього виходів:** {len(self.mechanisms)}")
        lines.append(f"- **Інвертованих сигналів:** {sum(1 for m in self.mechanisms for key in ['DI_Speed_Invert', 'DI_Breaker_Invert', 'DI_Overflow_Invert', 'DO_Run_Invert'] if m[key].upper() in ('TRUE', '1', 'YES'))}")
        
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))
        
        print(f"✓ Згенеровано: {output_file}")
    
    def generate_all(self, output_dir: str = "."):
        """Згенерувати всі файли з валідацією"""
        print("\n" + "="*80)
        print("HAL CODE GENERATOR (Excel Edition)")
        print("="*80)
        print()
        
        # Валідація
        errors = self.validate_config()
        if errors:
            print("❌ ПОМИЛКИ ВАЛІДАЦІЇ:")
            for error in errors:
                print(f"  ✗ {error}")
            print()
            return False
        
        print("✓ Валідація пройшла успішно")
        print()
        
        # Створити директорію
        os.makedirs(output_dir, exist_ok=True)
        
        # Генерація
        self.generate_hal_read(os.path.join(output_dir, "FC_HAL_Read_Redler_Static.scl"))
        self.generate_hal_write(os.path.join(output_dir, "FC_HAL_Write_Redler_Static.scl"))
        self.generate_db_io_config(os.path.join(output_dir, "DB_IO_Config_Doc.scl"))
        self.generate_documentation(os.path.join(output_dir, "HAL_IO_Mapping.md"))
        
        print()
        print("="*80)
        print("✓ ГЕНЕРАЦІЯ ЗАВЕРШЕНА")
        print("="*80)
        print(f"Файли збережено у: {os.path.abspath(output_dir)}")
        print()
        print("📋 Імпортуйте у TIA Portal:")
        print("  1. FC_HAL_Read_Redler_Static.scl")
        print("  2. FC_HAL_Write_Redler_Static.scl")
        print("  3. DB_IO_Config_Doc.scl")
        print()
        
        return True


def main():
    """Головна функція"""
    parser = argparse.ArgumentParser(
        description='Генератор HAL коду для TIA Portal з Excel/CSV конфігурації',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Приклади використання:
  python generate_hal_excel.py IO_Config.xlsx
  python generate_hal_excel.py IO_Config.xls
  python generate_hal_excel.py IO_Config.csv -o ./generated
  
Підтримувані формати:
  - .xlsx (Office 2007+) - потрібен openpyxl
  - .xls  (Office 97-2003) - потрібен xlrd
  - .csv  (текстовий) - вбудована підтримка

Встановлення залежностей:
  pip install openpyxl xlrd
        """
    )
    
    parser.add_argument('excel_file', 
                       help='Шлях до Excel/CSV файлу конфігурації')
    parser.add_argument('-o', '--output', 
                       default='.', 
                       help='Директорія для збереження файлів (за замовчуванням: поточна)')
    parser.add_argument('-v', '--verbose', 
                       action='store_true',
                       help='Детальний вивід')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.excel_file):
        print(f"❌ Файл не знайдено: {args.excel_file}")
        return 1
    
    try:
        generator = HALGenerator(args.excel_file)
        
        if generator.generate_all(args.output):
            return 0
        else:
            return 1
            
    except Exception as e:
        print(f"\n❌ КРИТИЧНА ПОМИЛКА:")
        print(f"   {type(e).__name__}: {e}")
        
        if args.verbose:
            import traceback
            print("\nДетальна інформація:")
            traceback.print_exc()
        
        return 1


if __name__ == '__main__':
    sys.exit(main())